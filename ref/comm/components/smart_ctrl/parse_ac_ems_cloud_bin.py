"""
@File      : parse_ac_ems_cloud_bin.py
@Version   : 1.0
@Author    : GitHub Copilot
@Date      : 2025/11/27
@Brief     : 解析 ac_ems_cloud_ctrl_config_t 二进制或16进制文本文件为结构化可读内容
@Details   :
  - 输入: ac_ems_cloud_ctrl_config_t 结构体原始二进制文件 或 4000字符16进制文本文件
  - 输出: 结构化文本（控制台打印或导出为CSV/Excel）
"""

import struct
import sys
import os
import csv
import binascii

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ===================== 结构体参数（需与C端保持一致） =====================
AC_EMS_CLOUD_REG_MAX_LEN = 1000
AC_EMS_CLOUD_UNIT_MAX_CNT = 50
AC_EMS_CLOUD_CHANNEL_MAX = 6
TIME_UNIT_MAX_LEN = 10
AC_EMS_CLOUD_UNIT_MAX_LEN = TIME_UNIT_MAX_LEN + (AC_EMS_CLOUD_CHANNEL_MAX * 4)
AC_EMS_CLOUD_CTRL_CONFIG_SIZE = AC_EMS_CLOUD_REG_MAX_LEN * 2

CSV_HEADERS = [
    "序号", "触发时间", "小时", "分钟", "秒", "星期", "日期", "月份",
    "时间类型",
    "reg_type", "device_id"
] + [f"通道{i+1}功率(W)" for i in range(AC_EMS_CLOUD_CHANNEL_MAX)]

WEEK_NAMES = ["周日", "周一", "周二", "周三", "周四", "周五", "周六"]

def parse_week_sel(week_sel):
    return "、".join([WEEK_NAMES[i] for i in range(7) if (week_sel >> i) & 1]) or "无"

def parse_month_sel(month_sel):
    return "、".join([f"{i+1}月" for i in range(12) if (month_sel >> i) & 1]) or "无"

def parse_day_sel(day_sel):
    return "、".join([f"{i+1}日" for i in range(31) if (day_sel >> i) & 1]) or "无"

def load_bin_or_hex(filename):
    """
    支持输入二进制文件或4000字符16进制文本文件，返回2000字节二进制数据
    """
    if not os.path.exists(filename):
        print(f"\n错误: 文件 '{filename}' 不存在或路径不正确。")
        return None

    with open(filename, "rb") as f:
        raw = f.read()

    # 判断是否为16进制文本（ASCII且长度接近4000）
    try:
        text = raw.decode("utf-8")
        hex_string = "".join(text.split())
        if len(hex_string) == 4000:
            try:
                binary_data = binascii.unhexlify(hex_string)
            except (binascii.Error, ValueError) as e:
                print(f"错误: 16进制字符串格式不正确: {e}")
                return None
            if len(binary_data) != 2000:
                print(f"错误: 转换后字节数为{len(binary_data)}，应为2000字节！")
                return None
            print(f"已识别为16进制文本，成功转换为2000字节二进制数据。")
            return binary_data
    except UnicodeDecodeError:
        pass

    # 否则直接判断为二进制
    if len(raw) == 2000:
        print(f"已识别为2000字节二进制文件。")
        return raw
    else:
        print(f"错误: 文件长度为{len(raw)}字节，既不是2000字节二进制，也不是4000字符16进制文本。")
        return None

def parse_time_type(t):
    """时间类型解析（与 C 枚举对应）"""
    return {
        0: "标准时间点设置",
        1: "标准时间段设置",
        2: "预留2",
        3: "预留3"
    }.get(t, f"未知({t})")

def parse_tou_time_config(raw):
    # raw: bytes-like, 长度10
    val = int.from_bytes(raw, "little")
    time_sec   = (val >> 0)  & 0x1FFFF         # bit0-16: 一天内的秒(0~86399)
    week_sel   = (val >> 17) & 0x7F            # bit17-23: 周选择
    day_sel    = (val >> 24) & 0x7FFFFFFF      # bit24-54: 日选择 (31 bits)
    month_sel  = (val >> 55) & 0xFFF           # bit55-66: 月选择 (12 bits)
    reg_type   = (val >> 67) & 0x3             # bit67-68: reg_type (2 bits)
    device_id  = (val >> 69) & 0xFF            # bit69-76: device_id (8 bits)
    time_type  = (val >> 77) & 0x7             # bit77-79: time_type (3 bits)
    return time_sec, week_sel, day_sel, month_sel, reg_type, device_id, time_type

# 在 parse_ac_ems_cloud_ctrl_config_from_bytes 中调用 parse_time_type 并包含时间类型字段
def parse_ac_ems_cloud_ctrl_config_from_bytes(data):
    """
    解析 ac_ems_cloud_ctrl_config_t 二进制数据，返回计划信息和单元列表
    """
    if len(data) != AC_EMS_CLOUD_CTRL_CONFIG_SIZE:
        print(f"数据长度不符，期望{AC_EMS_CLOUD_CTRL_CONFIG_SIZE}字节，实际{len(data)}字节")
        return None, None

    # 1. 解析 ac_ems_cloud_tou_plan_ch_cnt_t
    ch_cnt_val, = struct.unpack("<H", data[:2])
    ch_cnt = ch_cnt_val & 0xFF
    time_num = (ch_cnt_val >> 8) & 0xFF

    # 2. 解析 ac_ems_cloud_tou_plan_ch_en_t
    ch_en_val, = struct.unpack("<H", data[2:4])
    diff_type = (ch_en_val >> 12) & 0xF

    plan_info = {
        "ch_cnt": ch_cnt,
        "time_num": time_num,
        "diff_type": diff_type
    }

    # 3. 解析 ac_ems_cloud_unit_config_t[]
    units = []
    offset = 4
    for i in range(AC_EMS_CLOUD_UNIT_MAX_CNT):
        unit_raw = data[offset:offset+AC_EMS_CLOUD_UNIT_MAX_LEN]
        if len(unit_raw) < AC_EMS_CLOUD_UNIT_MAX_LEN:
            break

        # 3.1 解析 tou_time_config_t
        time_sec, week_sel, day_sel, month_sel, reg_type, device_id, time_type = parse_tou_time_config(unit_raw[:TIME_UNIT_MAX_LEN])

        # 3.2 解析 channel_power
        channel_power = []
        for j in range(AC_EMS_CLOUD_CHANNEL_MAX):
            ch_offset = TIME_UNIT_MAX_LEN + j * 4
            ch_val = struct.unpack("<i", unit_raw[ch_offset:ch_offset+4])[0]
            channel_power.append(ch_val)

        units.append({
            "序号": i,
            "触发时间": f"{time_sec//3600:02}:{(time_sec%3600)//60:02}:{time_sec%60:02}",
            "小时": time_sec // 3600,
            "分钟": (time_sec % 3600) // 60,
            "秒": time_sec % 60,
            "星期": parse_week_sel(week_sel),
            "日期": parse_day_sel(day_sel),
            "月份": parse_month_sel(month_sel),
            "时间类型": parse_time_type(time_type),
            "reg_type": reg_type,
            "device_id": device_id,
            **{f"通道{j+1}功率(W)": channel_power[j] for j in range(AC_EMS_CLOUD_CHANNEL_MAX)}
        })
        offset += AC_EMS_CLOUD_UNIT_MAX_LEN

    return plan_info, units

def export_to_txt_csv(plan_info, units, output_filename):
    """将解析出的数据导出为 TXT (CSV 格式) 文件。"""
    with open(output_filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["--- AC_EMS Cloud TOU计划信息 ---"])
        for key, value in plan_info.items():
            writer.writerow([key, value])
        writer.writerow([])
        writer.writerow(["--- AC_EMS Cloud TOU单元配置 ---"])
        writer.writerow(CSV_HEADERS)
        for u in units[:plan_info["time_num"]]:
            writer.writerow([u[k] for k in CSV_HEADERS])
    print(f"\n结果已成功导出到: {output_filename}")

def export_to_excel(plan_info, units, output_filename):
    """将解析出的数据导出为格式化的 Excel (.xlsx) 文件。"""
    if not OPENPYXL_AVAILABLE:
        print("\n错误: 'openpyxl' 库未安装。无法导出为 Excel 文件。")
        print("请运行 'pip install openpyxl' 来安装。")
        return

    wb = openpyxl.Workbook()
    ws_plan = wb.active
    ws_plan.title = "TOU计划信息"
    ws_plan.append(["项目", "值"])
    for key, value in plan_info.items():
        ws_plan.append([key, value])

    ws_units = wb.create_sheet(title="TOU单元配置")
    ws_units.append(CSV_HEADERS)
    bold_font = Font(bold=True)
    for cell in ws_units[1]:
        cell.font = bold_font
    for u in units[:plan_info["time_num"]]:
        ws_units.append([u[k] for k in CSV_HEADERS])

    # 自动调整列宽
    for ws in [ws_plan, ws_units]:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_filename)
    print(f"\n结果已成功导出到: {output_filename}")

if __name__ == "__main__":
    try:
        input_filename = input("请输入要解析的 ac_ems_cloud_ctrl_config 文件名 (支持 .bin 或 16进制.txt): ").strip()
        data = load_bin_or_hex(input_filename)
        if data:
            plan_info, units = parse_ac_ems_cloud_ctrl_config_from_bytes(data)
            if plan_info and units:
                while True:
                    print("\n请选择输出格式:")
                    print("  1: TXT (CSV 格式)")
                    print("  2: Excel (.xlsx)")
                    choice = input("请输入选项 (1 或 2): ")
                    base_filename = os.path.splitext(input_filename)[0]
                    if choice == '1':
                        output_filename = f"{base_filename}_parsed.txt"
                        export_to_txt_csv(plan_info, units, output_filename)
                        break
                    elif choice == '2':
                        output_filename = f"{base_filename}_parsed.xlsx"
                        export_to_excel(plan_info, units, output_filename)
                        break
                    else:
                        print("无效输入，请输入 1 或 2。")
    except KeyboardInterrupt:
        print("\n\n操作已由用户取消。")
    except Exception as e:
        print(f"\n发生未知错误: {e}")