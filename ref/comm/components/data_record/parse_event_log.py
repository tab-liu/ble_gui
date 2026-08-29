"""
@File      : parse_event_log.py
@Version   : 1.1
@Author    : GitHub Copilot
@Date      : 2025/10/24
@Brief     : 用于解析 ESP32 项目中 EventHistoryData_Struct 结构生成的日志文件。
@Details   :
  - 输入: 一个 .txt 文件，其内容为由 EventHistoryData_Struct 结构体
          连续组成的二进制日志数据的十六进制字符串表示。
  - 约束: 文件不包含文件头，直接从第一条记录开始。
  - 处理:
    1. 读取文件中的十六进制字符串。
    2. 将其转换回原始的二进制字节流。
    3. 逐条解析每个20字节的日志记录。
    4. 对位域（如故障/告警状态）和告警码进行解析和描述映射。
    5. 将解析结果进行格式化，使其易于人类阅读。
  - 输出:
    - 在控制台打印解析结果。
    - 可选择将结果导出为结构化的 TXT (CSV格式) 或 Excel (.xlsx) 文件。
"""

# ==============================================================================
# 依赖库导入
# ==============================================================================
import struct         # 用于处理二进制数据
import datetime       # 用于转换时间戳
import sys            # 用于处理程序退出
import os             # 用于处理文件路径
import csv            # 用于生成 CSV 文件
import binascii       # 用于十六进制字符串和二进制数据转换

# 尝试导入 openpyxl，如果失败则标记为不可用
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==============================================================================
# 全局常量定义 (与 C 结构体 EventHistoryData_Struct 严格对应)
# ==============================================================================

# 定义每条记录的固定长度（字节）
EVENT_RECORD_LEN = 20

# 定义二进制解包格式字符串
# '<' 表示小端字节序 (Little-endian)
RECORD_STRUCT_FORMAT = '<IBHHBBBBBHH'

# 状态枚举的可读映射
STATE_MAP = {
    1: "发生 (Happen)",
    0: "清除 (Clear)"
}

# 新增：告警码的可读映射
WARN_CODE_MAP = {
    15000: "系统开机(MCU上电)",
    15001: "系统关机(MCU下电)",
    15002: "定时记录",
    15003: "SPI flash故障",
    15004: "EEPROM故障",
    15005: "Sub 1GHz(CC1312)通信超时",
    15006: "4G模块通信超时",
    15007: "4G SIM卡未插",
    15008: "USB输出接口供电保护",
    15009: "ARM(IOT板内)通信超时",
    15010: "RTC(IOT板内)通信故障",
    15011: "meter(BL6552)1通信超时",
    15012: "meter(BL6552)2通信超时",
    15013: "meter(BL6552)3通信超时",
    15014: "meter(BL6552)4通信超时",
    15015: "meter(BL6552)5通信超时",
}

# 定义导出文件的列标题
CSV_HEADERS = [
    "记录序号", "可读时间", "协议版本", 
    "故障码", "故障状态", 
    "告警码", "告警状态", "告警描述",
    "IOT数量", "逆变器数量", "PACK数量", "WiFi节点数", "Sub1G节点数",
    "WiFi连接时长(分)", "Sub1G连接时长(分)", "原始时间戳"
]

# ==============================================================================
# 核心功能函数
# ==============================================================================

def parse_event_log_file(filepath):
    """
    解析包含事件日志十六进制字符串的文件。

    @param  filepath: 日志文件的路径。
    @return: 一个包含解析后记录的列表，如果失败则返回 None。
    """
    if not os.path.exists(filepath):
        print(f"\n错误: 文件 '{filepath}' 不存在或路径不正确。")
        return None

    # 1. 读取文件内容
    with open(filepath, 'r') as f:
        hex_string = f.read().strip()

    # 2. 将十六进制字符串转换为二进制数据
    try:
        hex_string = "".join(hex_string.split())
        if not hex_string:
            print("\n错误: 文件内容为空。")
            return None
        binary_content = binascii.unhexlify(hex_string)
    except (binascii.Error, TypeError) as e:
        print(f"\n错误: 文件内容不是有效的十六进制字符串。 {e}")
        return None

    file_size = len(binary_content)
    print(f"\n--- 开始解析文件: '{filepath}' (有效数据大小: {file_size} 字节) ---")

    # 3. 验证文件大小是否为记录长度的整数倍
    if file_size % EVENT_RECORD_LEN != 0:
        print(f"警告: 文件大小 ({file_size}字节) 不是记录长度 ({EVENT_RECORD_LEN}字节) 的整数倍。")
        print("末尾不完整的数据将被忽略。")

    # 4. 循环解析所有日志记录
    num_records_in_file = file_size // EVENT_RECORD_LEN
    parsed_records = []

    print(f"\n--- 发现 {num_records_in_file} 条日志记录 ---")

    for i in range(num_records_in_file):
        record_chunk = binary_content[i * EVENT_RECORD_LEN : (i + 1) * EVENT_RECORD_LEN]
        
        # 使用一个完整的格式字符串一次性解包所有11个字段
        (timestamp, ver_protocol, fault_raw, warn_raw, num_iot, num_inv, num_pack, 
         num_udp_net, num_sub1g_net, time_wifi_connect, time_sub1g_connect) = \
         struct.unpack_from('<IBHHBBBBBHH', record_chunk)

        # 解析位域
        fault_code = fault_raw & 0x7FFF
        fault_state = (fault_raw >> 15) & 0x01
        warn_code = warn_raw & 0x7FFF
        warn_state = (warn_raw >> 15) & 0x01

        # 新增：解析告警码描述
        warn_description = WARN_CODE_MAP.get(warn_code, f"未定义告警({warn_code})")

        # 格式化时间戳
        try:
            dt_object = datetime.datetime.fromtimestamp(timestamp)
            time_str = dt_object.strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, OSError):
            time_str = "无效时间戳"

        # 将解析出的数据存入字典
        record_dict = {
            "记录序号": i + 1,
            "可读时间": time_str,
            "协议版本": ver_protocol,
            "故障码": fault_code,
            "故障状态": STATE_MAP.get(fault_state, f"未知({fault_state})"),
            "告警码": warn_code,
            "告警状态": STATE_MAP.get(warn_state, f"未知({warn_state})"),
            "告警描述": warn_description,
            "IOT数量": num_iot,
            "逆变器数量": num_inv,
            "PACK数量": num_pack,
            "WiFi节点数": num_udp_net,
            "Sub1G节点数": num_sub1g_net,
            "WiFi连接时长(分)": time_wifi_connect,
            "Sub1G连接时长(分)": time_sub1g_connect,
            "原始时间戳": timestamp
        }
        parsed_records.append(record_dict)

    print("\n--- 解析完成 ---")
    return parsed_records

def export_to_txt_csv(records, output_filename):
    """将解析出的数据导出为 TXT (CSV 格式) 文件。"""
    with open(output_filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(CSV_HEADERS)
        for record in records:
            writer.writerow(record.values())
    print(f"\n结果已成功导出到: {output_filename}")

def export_to_excel(records, output_filename):
    """将解析出的数据导出为格式化的 Excel (.xlsx) 文件。"""
    if not OPENPYXL_AVAILABLE:
        print("\n错误: 'openpyxl' 库未安装。无法导出为 Excel 文件。")
        print("请运行 'pip install openpyxl' 来安装。")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "事件日志记录"
    ws.append(CSV_HEADERS)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for record in records:
        ws.append(list(record.values()))

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_filename)
    print(f"\n结果已成功导出到: {output_filename}")

# ==============================================================================
# 主程序入口
# ==============================================================================

if __name__ == "__main__":
    try:
        input_filename = input("请输入要解析的事件日志文件名 (例如: event_log_dump.txt): ")
        
        records_list = parse_event_log_file(input_filename)

        if records_list:
            # 打印前5条记录作为预览
            print("\n--- 解析结果预览 (前5条) ---")
            for i, record in enumerate(records_list[:5]):
                print(f"记录 #{i+1}: {record['可读时间']}, 告警: {record['告警描述']}, 状态: {record['告警状态']}")
            print("-" * 30)

            while True:
                print("\n请选择输出格式:")
                print("  1: TXT (CSV 格式)")
                print("  2: Excel (.xlsx)")
                choice = input("请输入选项 (1 或 2，或直接回车跳过): ")
                
                base_filename = os.path.splitext(input_filename)[0]

                if choice == '1':
                    output_filename = f"{base_filename}_parsed_events.txt"
                    export_to_txt_csv(records_list, output_filename)
                    break
                elif choice == '2':
                    output_filename = f"{base_filename}_parsed_events.xlsx"
                    export_to_excel(records_list, output_filename)
                    break
                elif choice == '':
                    print("\n未选择导出，程序结束。")
                    break
                else:
                    print("无效输入，请输入 1, 2 或直接回车。")

    except KeyboardInterrupt:
        print("\n\n操作已由用户取消。")
    except Exception as e:
        print(f"\n发生未知错误: {e}")