"""
@File      : parse_mqtt_log.py
@Version   : 1.4
@Author    : lixingyu
@Date      : 2026/8/7
@Brief     : 用于解析 ESP32 项目中 mqtt_log 模块生成的日志文件。
@Details   :
  - 输入: 一个 .txt 文件，其内容为从设备读取的二进制日志数据的十六进制字符串表示。
  - 处理:
    1. 读取文件中的十六进制字符串。
    2. 将其转换回原始的二进制字节流。
    3. 根据 C 头文件 (mqtt_log.h) 中定义的结构体，解析文件头和每一条日志记录。
    4. 将解析结果进行格式化，使其易于人类阅读。
  - 输出:
    - 在控制台打印解析结果。
    - 可选择将结果导出为结构化的 TXT (CSV格式) 或 Excel (.xlsx) 文件。
      （Excel 导出时所有单元格内容居中对齐）
"""

# ==============================================================================
# 依赖库导入
# ==============================================================================
import struct
import datetime
import sys
import os
import csv
import binascii

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==============================================================================
# 全局常量定义 (与 C 头文件 mqtt_log.h 严格对应)
# ==============================================================================

# 定义文件头和每条记录的固定长度（字节）
MQTT_LOG_FILE_HEADER_LEN = 64
MQTT_LOG_LEN = 64

# 定义二进制解包格式字符串
# '<' 表示小端字节序 (Little-endian)
# 对应 mqtt_log_file_header_t 中前 8 字节有效字段
HEADER_STRUCT_FORMAT = '<HHHH'

# 对应 mqtt_conn_fail_log_t 结构体（packed）前 56 字节有效数据
# 字段顺序严格按 C 定义：
#   log_id: uint32_t
#   fail_timestamp: uint64_t
#   svc_timestamp: uint64_t
#   uptime_seconds: uint32_t
#   use_mqtts: uint8_t
#   password: uint64_t
#   reason: uint8_t
#   sys_errno: int32_t
#   error_type: uint8_t
#   connect_return_code: uint8_t
#   network_type: uint8_t
#   rssi: int8_t
#   retry_delay_ms: uint16_t
#   esp_tls_last_esp_err: int32_t
#   esp_tls_stack_err: int32_t
#   esp_tls_cert_verify_flags: int32_t
#   reserved: 8 bytes
RECORD_STRUCT_FORMAT = '< I Q Q I B Q B i B B B b H i i i 8s'

# ==============================================================================
# 枚举值到字符串的映射
# ==============================================================================

MqttFailReason_MAP = {
    0: "未知",
    1: "DNS解析失败",
    2: "TCP连接失败",
    3: "TCP获取时间失败",
    4: "证书更新下载失败",
    5: "MQTT登录过程失败",
    6: "MQTT拒绝(非0原因码)",
    7: "MQTT订阅失败",
    8: "网络切换导致中断",
    9: "协议层关闭或客户端主动断开",
    10: "设备主动终止(状态或参数更新)",
    11: "应用层等待超时(poll 超时)",
    12: "等待 CONNACK 时对端关闭或超时",
}

ESP_MQTT_ERROR_TYPE_MAP = {
    0: "无错误",
    1: "TCP传输错误",
    2: "连接被拒绝",
    3: "订阅失败",
}

MQTT_CONNECTION_CODE_MAP = {
    0: "连接接受",
    1: "协议错误",
    2: "ID拒绝",
    3: "服务器不可用",
    4: "用户名错误",
    5: "未授权",
}

NETWORK_TYPE_MAP = {
    0: "Ethernet",
    1: "Wi-Fi",
    2: "PPP",
}

# ==============================================================================
# POSIX / ESP32 errno 错误码表
# ==============================================================================
ERRNO_TABLE = {
    0:   ("OK", "Success"),
    1:   ("EPERM", "Operation not permitted"),
    2:   ("ENOENT", "No such file or directory"),
    3:   ("ESRCH", "No such process"),
    4:   ("EINTR", "Interrupted system call"),
    5:   ("EIO", "I/O error"),
    6:   ("ENXIO", "No such device or address"),
    7:   ("E2BIG", "Argument list too long"),
    8:   ("ENOEXEC", "Exec format error"),
    9:   ("EBADF", "Bad file descriptor"),
    10:  ("ECHILD", "No child processes"),
    11:  ("EAGAIN", "Try again"),
    12:  ("ENOMEM", "Out of memory"),
    13:  ("EACCES", "Permission denied"),
    14:  ("EFAULT", "Bad address"),
    15:  ("ENOTBLK", "Block device required"),
    16:  ("EBUSY", "Device or resource busy"),
    17:  ("EEXIST", "File exists"),
    18:  ("EXDEV", "Cross-device link"),
    19:  ("ENODEV", "No such device"),
    20:  ("ENOTDIR", "Not a directory"),
    21:  ("EISDIR", "Is a directory"),
    22:  ("EINVAL", "Invalid argument"),
    23:  ("ENFILE", "File table overflow"),
    24:  ("EMFILE", "Too many open files"),
    25:  ("ENOTTY", "Not a typewriter"),
    26:  ("ETXTBSY", "Text file busy"),
    27:  ("EFBIG", "File too large"),
    28:  ("ENOSPC", "No space left on device"),
    29:  ("ESPIPE", "Illegal seek"),
    30:  ("EROFS", "Read-only file system"),
    31:  ("EMLINK", "Too many links"),
    32:  ("EPIPE", "Broken pipe"),
    33:  ("EDOM", "Math argument out of domain"),
    34:  ("ERANGE", "Math result not representable"),
    88:  ("ENOTSOCK", "Socket operation on non-socket"),
    89:  ("EDESTADDRREQ", "Destination address required"),
    90:  ("EMSGSIZE", "Message too long"),
    91:  ("EPROTOTYPE", "Protocol wrong type for socket"),
    92:  ("ENOPROTOOPT", "Protocol not available"),
    93:  ("EPROTONOSUPPORT", "Protocol not supported"),
    94:  ("ESOCKTNOSUPPORT", "Socket type not supported"),
    95:  ("EOPNOTSUPP", "Operation not supported on transport endpoint"),
    96:  ("EPFNOSUPPORT", "Protocol family not supported"),
    97:  ("EAFNOSUPPORT", "Address family not supported by protocol"),
    98:  ("EADDRINUSE", "Address already in use"),
    99:  ("EADDRNOTAVAIL", "Cannot assign requested address"),
    100: ("ENETDOWN", "Network is down"),
    101: ("ENETUNREACH", "Network is unreachable"),
    102: ("ENETRESET", "Network dropped connection because of reset"),
    103: ("ECONNABORTED", "Software caused connection abort"),
    104: ("ECONNRESET", "Connection reset by peer"),
    105: ("ENOBUFS", "No buffer space available"),
    106: ("EISCONN", "Transport endpoint is already connected"),
    107: ("ENOTCONN", "Transport endpoint is not connected"),
    108: ("ESHUTDOWN", "Cannot send after transport endpoint shutdown"),
    109: ("ETOOMANYREFS", "Too many references: cannot splice"),
    110: ("ETIMEDOUT", "Connection timed out"),
    111: ("ECONNREFUSED", "Connection refused"),
    112: ("EHOSTDOWN", "Host is down"),
    113: ("EHOSTUNREACH", "No route to host"),
    114: ("EALREADY", "Operation already in progress"),
    115: ("EINPROGRESS", "Operation now in progress"),
    116: ("ESTALE", "Stale file handle"),
}

def format_errno(code: int) -> str:
    if code in ERRNO_TABLE:
        return f"{ERRNO_TABLE[code][0]} ({code}): {ERRNO_TABLE[code][1]}"
    return f"UNKNOWN ({code})"

# ==============================================================================
# TLS 错误码映射 (基于 ESP-IDF 官方定义)
# ESP_ERR_ESP_TLS_BASE = 0x8000
# ==============================================================================

ESP_TLS_ERR_MAP = {
    0x0: "ESP_OK (成功)",
    0xFFFFFFFF: "ESP_FAIL (通用失败)",
    0x101: "ESP_ERR_NO_MEM (内存不足)",
    0x102: "ESP_ERR_INVALID_ARG (参数无效)",
    0x103: "ESP_ERR_INVALID_STATE (状态无效)",
    0x104: "ESP_ERR_INVALID_SIZE (大小无效)",
    0x105: "ESP_ERR_NOT_FOUND (未找到)",
    0x106: "ESP_ERR_NOT_SUPPORTED (不支持)",
    0x107: "ESP_ERR_TIMEOUT (超时)",
    0x108: "ESP_ERR_INVALID_RESPONSE (无效响应)",
    0x109: "ESP_ERR_INVALID_CRC (CRC 错误)",
    0x10A: "ESP_ERR_INVALID_VERSION (版本无效)",
    0x10B: "ESP_ERR_INVALID_MAC (MAC 无效)",
    0x3000: "ESP_ERR_WIFI_BASE (Wi-Fi 基础错误)",
    0x8000: "ESP_ERR_ESP_TLS_BASE (ESP-TLS 基础错误码)",
    0x8001: "ESP_ERR_ESP_TLS_CANNOT_RESOLVE_HOSTNAME (主机名解析失败)",
    0x8002: "ESP_ERR_ESP_TLS_CANNOT_CREATE_SOCKET (创建 Socket 失败)",
    0x8003: "ESP_ERR_ESP_TLS_UNSUPPORTED_PROTOCOL_FAMILY (不支持协议族)",
    0x8004: "ESP_ERR_ESP_TLS_FAILED_CONNECT_TO_HOST (连接主机失败)",
    0x8005: "ESP_ERR_ESP_TLS_SOCKET_SETOPT_FAILED (设置 Socket 选项失败)",
    0x8006: "ESP_ERR_ESP_TLS_CONNECTION_TIMEOUT (连接超时)",
    0x8007: "ESP_ERR_ESP_TLS_SE_FAILED (Secure Element 错误)",
    0x8008: "ESP_ERR_ESP_TLS_TCP_CLOSED_FIN (TCP 连接被关闭)",
    0x8009: "ESP_ERR_ESP_TLS_SERVER_HANDSHAKE_TIMEOUT (服务器握手超时)",
    0x8010: "ESP_ERR_MBEDTLS_CERT_PARTLY_OK (证书解析部分成功)",
    0x8011: "ESP_ERR_MBEDTLS_CTR_DRBG_SEED_FAILED (随机数生成失败)",
    0x8012: "ESP_ERR_MBEDTLS_SSL_SET_HOSTNAME_FAILED (设置主机名失败)",
    0x8013: "ESP_ERR_MBEDTLS_SSL_CONFIG_DEFAULTS_FAILED (配置默认值失败)",
    0x8014: "ESP_ERR_MBEDTLS_SSL_CONF_ALPN_PROTOCOLS_FAILED (配置 ALPN 协议失败)",
    0x8015: "ESP_ERR_MBEDTLS_X509_CRT_PARSE_FAILED (X.509 证书解析失败)",
    0x8016: "ESP_ERR_MBEDTLS_SSL_CONF_OWN_CERT_FAILED (配置证书失败)",
    0x8017: "ESP_ERR_MBEDTLS_SSL_SETUP_FAILED (TLS 设置失败)",
    0x8018: "ESP_ERR_MBEDTLS_SSL_WRITE_FAILED (写入失败)",
    0x8019: "ESP_ERR_MBEDTLS_PK_PARSE_KEY_FAILED (解析私钥失败)",
    0x801A: "ESP_ERR_MBEDTLS_SSL_HANDSHAKE_FAILED (TLS 握手失败)",
    0x801B: "ESP_ERR_MBEDTLS_SSL_CONF_PSK_FAILED (配置 PSK 失败)",
    0x801C: "ESP_ERR_MBEDTLS_SSL_TICKET_SETUP_FAILED (会话票证设置失败)",
    0x801D: "ESP_ERR_MBEDTLS_SSL_READ_FAILED (读取失败)",
}

MBEDTLS_SSL_ERR_MAP = {
    0x7280: "MBEDTLS_ERR_SSL_CONN_EOF (对端意外关闭连接)",
    0x6800: "MBEDTLS_ERR_SSL_TIMEOUT (操作超时)",
    0x6900: "MBEDTLS_ERR_SSL_WANT_READ (需要等待可读)",
    0x6880: "MBEDTLS_ERR_SSL_WANT_WRITE (需要等待可写)",
    0x7880: "MBEDTLS_ERR_SSL_PEER_CLOSE_NOTIFY (对端发送关闭通知)",
    0x7100: "MBEDTLS_ERR_SSL_BAD_INPUT_DATA (输入数据无效)",
    0x7180: "MBEDTLS_ERR_SSL_INVALID_MAC (消息认证码校验失败)",
    0x7200: "MBEDTLS_ERR_SSL_INVALID_RECORD (记录层数据无效)",
    0x7080: "MBEDTLS_ERR_SSL_FEATURE_UNAVAILABLE (请求功能不可用)",
    0x6C00: "MBEDTLS_ERR_SSL_UNEXPECTED_MESSAGE (收到意外握手消息)",
    0x6E80: "MBEDTLS_ERR_SSL_HANDSHAKE_FAILURE (握手失败)",
    0x6C80: "MBEDTLS_ERR_SSL_NO_CIPHER_CHOSEN (没有可用的加密套件)",
    0x6B80: "MBEDTLS_ERR_SSL_CA_CHAIN_REQUIRED (需要 CA 证书链)",
    0x7000: "MBEDTLS_ERR_SSL_PEER_VERIFY_FAILED (对端证书验证失败)",
    0x7400: "MBEDTLS_ERR_SSL_BAD_CERTIFICATE (证书格式错误)",
    0x7F00: "MBEDTLS_ERR_SSL_ALLOC_FAILED (内存分配失败)",
}

X509_CERT_FLAGS_MAP = {
    0x01: "证书已过期 (BADCERT_EXPIRED)",
    0x02: "证书已被吊销 (BADCERT_REVOKED)",
    0x04: "证书 CN 不匹配 (BADCERT_CN_MISMATCH)",
    0x08: "证书不被信任 (BADCERT_NOT_TRUSTED)",
    0x10: "CRL 不被信任 (BADCRL_NOT_TRUSTED)",
    0x20: "CRL 已过期 (BADCRL_EXPIRED)",
    0x40: "缺少证书 (BADCERT_MISSING)",
    0x80: "跳过了验证 (BADCERT_SKIP_VERIFY)",
    0x0100: "其他证书错误 (BADCERT_OTHER)",
    0x0200: "证书尚未生效 (BADCERT_FUTURE)",
    0x0400: "CRL 尚未生效 (BADCRL_FUTURE)",
    0x0800: "密钥用途不符 (BADCERT_KEY_USAGE)",
    0x1000: "扩展密钥用途不符 (BADCERT_EXT_KEY_USAGE)",
    0x2000: "命名空间证书类型不符 (BADCERT_NS_CERT_TYPE)",
}

def format_tls_error(esp_err, stack_err, cert_flags):
    """格式化 TLS 错误信息（值为0时不显示任何内容）"""
    errors = []
    
    # 格式化 ESP-TLS ESP-IDF 错误码
    if esp_err != 0:
        error_str = ESP_TLS_ERR_MAP.get(esp_err, f"UNKNOWN (0x{esp_err:08X})")
        errors.append(f"ESP-TLS: {error_str}")
    
    # 格式化底层 TLS 栈错误码
    if stack_err != 0:
        error_str = MBEDTLS_SSL_ERR_MAP.get(stack_err, f"MBEDTLS (0x{stack_err:08X})")
        errors.append(f"MBEDTLS: {error_str}")
    
    # 格式化证书验证标志位
    if cert_flags != 0:
        flags = []
        for flag_value, flag_name in X509_CERT_FLAGS_MAP.items():
            if cert_flags & flag_value:
                flags.append(flag_name)
        if flags:
            errors.append(f"证书验证: {', '.join(flags)}")
        else:
            errors.append(f"证书验证: 0x{cert_flags:04X}")
    
    return "\n".join(errors) if errors else ""

# ==============================================================================
# 定义导出文件的列标题
# ==============================================================================

CSV_HEADERS = [
    "日志序号", "日志ID", "失败时间戳", "可读时间",
    "服务器时间戳", "可读时间(服务器)", "运行时长(秒)", "是否MQTTS", 
    "密码", "失败原因", "系统错误码(详细)", "错误类型", 
    "连接拒绝码", "网络类型", "信号强度", "重连延迟(ms)",
    "ESP-TLS错误码", "MBEDTLS错误码", "证书验证标志值"
]

# ==============================================================================
# 核心功能函数
# ==============================================================================

def parse_mqtt_log_file(filepath):
    if not os.path.exists(filepath):
        print(f"\n错误: 文件 '{filepath}' 不存在或路径不正确。")
        return None, None
    if not os.path.isfile(filepath):
        print(f"\n错误: '{filepath}' 不是文件，而是一个目录。")
        return None, None

    with open(filepath, 'r') as f:
        hex_string = f.read().strip()

    try:
        hex_string = "".join(hex_string.split())
        binary_content = binascii.unhexlify(hex_string)
    except (binascii.Error, TypeError) as e:
        print(f"\n错误: 文件内容不是有效的十六进制字符串。 {e}")
        return None, None

    file_size = len(binary_content)
    print(f"\n--- 开始解析文件: '{filepath}' (有效数据大小: {file_size} 字节) ---")

    if file_size < MQTT_LOG_FILE_HEADER_LEN:
        print(f"错误: 数据大小 ({file_size}字节) 不足以包含一个完整的文件头 ({MQTT_LOG_FILE_HEADER_LEN}字节)。")
        return None, None

    # 解析文件头（只解析前 8 字节有效数据）
    header_data = binary_content[:8]
    record_ver, max_records, current_records, write_index = struct.unpack(HEADER_STRUCT_FORMAT, header_data)
    header_info = {
        "record_ver": record_ver,
        "max_records": max_records,
        "current_records": current_records,
        "write_index": write_index
    }
    print("\n--- 文件头信息 ---")
    for key, value in header_info.items():
        print(f"{key:<20}: {value}")
    print("-" * 20)

    # 解析记录区
    records_data = binary_content[MQTT_LOG_FILE_HEADER_LEN:]
    num_records_in_file = len(records_data) // MQTT_LOG_LEN
    parsed_records = []

    print(f"\n--- 发现 {num_records_in_file} 条日志记录 ---")

    for i in range(num_records_in_file):
        record_chunk = records_data[i * MQTT_LOG_LEN : (i + 1) * MQTT_LOG_LEN]
        
        # 使用修正后的格式字符串，传递完整的 64 字节记录数据
        fields = struct.unpack_from(RECORD_STRUCT_FORMAT, record_chunk[:64])
        (log_id, fail_timestamp, svc_timestamp, uptime_seconds,
         use_mqtts, password, reason, sys_errno,
         error_type, connect_return_code, network_type,
         rssi, retry_delay_ms, esp_tls_last_esp_err,
         esp_tls_stack_err, esp_tls_cert_verify_flags) = fields[:16]
        
        # 单独提取 reserved 字段
        reserved = record_chunk[56:64]

        # 格式化时间
        try:
            fail_dt_object = datetime.datetime.fromtimestamp(fail_timestamp)
            fail_time_str = fail_dt_object.strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, OSError):
            fail_time_str = "无效时间戳"

        try:
            svc_dt_object = datetime.datetime.fromtimestamp(svc_timestamp)
            svc_time_str = svc_dt_object.strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, OSError):
            svc_time_str = "无效时间戳"

        record_dict = {
            "日志序号": i + 1,
            "日志ID": log_id,
            "失败时间戳": fail_timestamp,
            "可读时间": fail_time_str,
            "服务器时间戳": svc_timestamp,
            "可读时间(服务器)": svc_time_str,
            "运行时长(秒)": uptime_seconds,
            "是否MQTTS": "是" if use_mqtts == 1 else "否",
            "密码": f"{password:08}",                # 固定8位显示，不足左侧补零
            "失败原因": MqttFailReason_MAP.get(reason, f"未知({reason})"),
            "系统错误码(详细)": format_errno(sys_errno),
            "错误类型": ESP_MQTT_ERROR_TYPE_MAP.get(error_type, f"未知({error_type})"),
            "连接拒绝码": MQTT_CONNECTION_CODE_MAP.get(connect_return_code, f"未知({connect_return_code})"),
            "网络类型": NETWORK_TYPE_MAP.get(network_type, f"未知({network_type})"),
            "信号强度": f"{rssi} dBm",
            "重连延迟(ms)": retry_delay_ms,
            "ESP-TLS错误码": format_tls_error(esp_tls_last_esp_err, 0, 0) or "无",
            "MBEDTLS错误码": format_tls_error(0, esp_tls_stack_err, 0) or "无",
            "证书验证标志值": format_tls_error(0, 0, esp_tls_cert_verify_flags) or "无"
        }
        parsed_records.append(record_dict)

    print("\n--- 解析完成 ---")
    return header_info, parsed_records

def export_to_txt_csv(header, records, output_filename):
    with open(output_filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["--- 文件头信息 ---"])
        for key, value in header.items():
            writer.writerow([key, value])
        writer.writerow([])
        writer.writerow(["--- 日志记录 ---"])
        writer.writerow(CSV_HEADERS)
        for record in records:
            writer.writerow(record.values())
    print(f"\n结果已成功导出到: {output_filename}")

def export_to_excel(header, records, output_filename):
    if not OPENPYXL_AVAILABLE:
        print("\n错误: 'openpyxl' 库未安装。无法导出为 Excel 文件。")
        print("请运行 'pip install openpyxl' 来安装。")
        return

    wb = openpyxl.Workbook()
    ws_header = wb.active
    ws_header.title = "文件头信息"
    ws_header.append(["项目", "值"])
    for key, value in header.items():
        ws_header.append([key, value])
    
    ws_records = wb.create_sheet(title="日志记录")
    ws_records.append(CSV_HEADERS)

    bold_font = Font(bold=True)
    for cell in ws_records[1]:
        cell.font = bold_font

    for record in records:
        ws_records.append(list(record.values()))

    # 设置列宽（19列：包括新增的TLS错误列）
    column_widths = {
        'A': 10, 'B': 10, 'C': 12, 'D': 20, 'E': 15,
        'F': 20, 'G': 15, 'H': 12, 'I': 12, 'J': 30,
        'K': 40, 'L': 20, 'M': 20, 'N': 20, 'O': 12,
        'P': 20, 'Q': 65, 'R': 60, 'S': 30
    }
    for i, _ in enumerate(CSV_HEADERS, start=1):
        col_letter = get_column_letter(i)
        ws_records.column_dimensions[col_letter].width = column_widths.get(col_letter, 15)

    ws_header.column_dimensions['A'].width = 30
    ws_header.column_dimensions['B'].width = 30

    # ================== 新增：所有单元格居中对齐 ==================
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 文件头工作表
    for row in ws_header.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    # 日志记录工作表
    for row in ws_records.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    wb.save(output_filename)
    print(f"\n结果已成功导出到: {output_filename}")

# ==============================================================================
# 主程序入口
# ==============================================================================

if __name__ == "__main__":
    try:
        input_filename = input("请输入要解析的日志文件名 (例如: mqtt_log_dump.txt): ")
        header_info, records_list = parse_mqtt_log_file(input_filename)
        if header_info and records_list:
            while True:
                print("\n请选择输出格式:")
                print("  1: TXT (CSV 格式)")
                print("  2: Excel (.xlsx)")
                choice = input("请输入选项 (1 或 2): ")
                base_filename = os.path.splitext(input_filename)[0]
                if choice == '1':
                    output_filename = f"{base_filename}_parsed.txt"
                    export_to_txt_csv(header_info, records_list, output_filename)
                    break
                elif choice == '2':
                    output_filename = f"{base_filename}_parsed.xlsx"
                    export_to_excel(header_info, records_list, output_filename)
                    break
                else:
                    print("无效输入，请输入 1 或 2。")
    except KeyboardInterrupt:
        print("\n\n操作已由用户取消。")
    except Exception as e:
        print(f"\n发生未知错误: {e}")