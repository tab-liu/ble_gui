"""
@File      : parse_reg_log.py
@Version   : 1.2 (最终版)
@Author    : GitHub Copilot
@Date      : 2025/09/26
@Brief     : 用于解析 ESP32 项目中 reg_change_log 模块生成的日志文件。
@Details   :
  - 输入: 一个 .txt 文件，其内容为从设备读取的二进制日志数据的十六进制字符串表示。
  - 处理:
    1. 读取文件中的十六进制字符串。
    2. 将其转换回原始的二进制字节流。
    3. 根据 C 头文件 (reg_change_log.h) 中定义的结构体，解析文件头和每一条日志记录。
    4. 将解析结果进行格式化，使其易于人类阅读。
  - 输出:
    - 在控制台打印解析结果。
    - 可选择将结果导出为结构化的 TXT (CSV格式) 或 Excel (.xlsx) 文件。
"""

# ==============================================================================
# 依赖库导入
# ==============================================================================
import struct         # 用于处理二进制数据，执行打包和解包操作
import datetime       # 用于将 UNIX 时间戳转换为可读的日期时间格式
import sys            # 用于处理程序退出
import os             # 用于处理文件路径和文件名
import csv            # 用于生成 CSV 格式的 TXT 文件
import binascii       # 用于在二进制和 ASCII 编码的十六进制字符串之间进行转换

# 尝试导入 openpyxl，如果失败则标记为不可用，以便在后续流程中优雅地处理
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==============================================================================
# 全局常量定义 (与 C 头文件 reg_change_log.h 严格对应)
# ==============================================================================

# 定义文件头和每条记录的固定长度（字节）
REG_CHANGE_FILE_HEADER_LEN = 20
REG_CHANGE_LOG_LEN = 20

# 定义二进制解包格式字符串
# '<' 表示小端字节序 (Little-endian)，这是大多数嵌入式系统的标准
HEADER_STRUCT_FORMAT = '<HHH'          # 对应 reg_change_file_header_t.s (uint16_t * 3 = 6字节)
RECORD_STRUCT_FORMAT = '<IBBHHHHBB'    # 对应 reg_param_change_record_t.s (16字节)

# 协议类型枚举的可读映射
PROTOCOL_TYPE_MAP = {
    1: "Modbus Alpha", 2: "Modbus Beta", 3: "Modbus Open", 4: "CAN Beta",
}

# 数据来源枚举的可读映射
DATA_SOURCE_MAP = {
    0: "Unknown", 1: "APP (BLE Broadcast)", 2: "APP (BLE Connection)",
    3: "WiFi (UDP)", 4: "WiFi (MQTT)", 5: "CAN Bus", 6: "UART",
    7: "Modbus TCP", 8: "IoT Internal",
}

# 定义导出文件的列标题
CSV_HEADERS = [
    "记录序号", "时间戳", "可读时间", "源地址", "目标地址",
    "起始地址/类型", "偏移地址", "原始值", "修改后值",
    "协议版本", "修改来源"
]

# ==============================================================================
# 核心功能函数
# ==============================================================================

def parse_log_file(filepath):
    """
    解析包含十六进制字符串的日志文件，并返回文件头信息和记录列表。

    @param  filepath: 日志文件的路径。
    @return: 一个元组 (header_info, parsed_records)，如果失败则返回 (None, None)。
    """
    if not os.path.exists(filepath):
        print(f"\n错误: 文件 '{filepath}' 不存在或路径不正确。")
        return None, None

    # 1. 读取文件内容
    # 以文本模式 ('r') 读取，因为文件内容是十六进制字符串，而不是原始二进制。
    with open(filepath, 'r') as f:
        hex_string = f.read().strip()

    # 2. 【关键步骤】将十六进制字符串转换为二进制数据
    try:
        # 移除字符串中所有可能的空格或换行符，以确保转换的纯粹性
        hex_string = "".join(hex_string.split())
        # 使用 binascii.unhexlify 将 '2C0107...' 这样的字符串转换为 b'\x2c\x01\x07...' 字节流
        binary_content = binascii.unhexlify(hex_string)
    except (binascii.Error, TypeError) as e:
        print(f"\n错误: 文件内容不是有效的十六进制字符串。 {e}")
        return None, None

    file_size = len(binary_content)
    print(f"\n--- 开始解析文件: '{filepath}' (有效数据大小: {file_size} 字节) ---")

    # 3. 验证并解析文件头
    if file_size < REG_CHANGE_FILE_HEADER_LEN:
        print(f"错误: 数据大小不足以包含一个完整的文件头。")
        return None, None

    header_data = binary_content[:REG_CHANGE_FILE_HEADER_LEN]
    # 从20字节的头数据块中，只解包前6个字节的结构化数据
    max_records, current_records, write_index = struct.unpack_from(HEADER_STRUCT_FORMAT, header_data)
    header_info = {
        "max_records": max_records,
        "current_records": current_records,
        "write_index": write_index
    }
    print("\n--- 文件头信息 ---")
    for key, value in header_info.items():
        print(f"{key:<20}: {value}")
    print("-" * 20)

    # 4. 循环解析所有日志记录
    records_data = binary_content[REG_CHANGE_FILE_HEADER_LEN:]
    num_records_in_file = len(records_data) // REG_CHANGE_LOG_LEN
    parsed_records = []

    print(f"\n--- 发现 {num_records_in_file} 条日志记录 ---")

    for i in range(num_records_in_file):
        record_chunk = records_data[i * REG_CHANGE_LOG_LEN : (i + 1) * REG_CHANGE_LOG_LEN]
        
        # 从20字节的记录数据块中，只解包前16个字节的结构化数据
        (timestamp, source_addr, target_addr, start_address, offset_address, 
         original_value, modified_value, protocol_version, modification_source) = \
         struct.unpack_from(RECORD_STRUCT_FORMAT, record_chunk)

        # 格式化时间戳
        try:
            dt_object = datetime.datetime.fromtimestamp(timestamp)
            time_str = dt_object.strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, OSError):
            time_str = "无效时间戳"

        # 将解析出的数据存入字典，便于后续处理和导出
        record_dict = {
            "记录序号": i + 1,
            "时间戳": timestamp,
            "可读时间": time_str,
            "源地址": source_addr,
            "目标地址": target_addr,
            "起始地址/类型": start_address,
            "偏移地址": offset_address,
            "原始值": original_value,
            "修改后值": modified_value,
            "协议版本": PROTOCOL_TYPE_MAP.get(protocol_version, f"未知({protocol_version})"),
            "修改来源": DATA_SOURCE_MAP.get(modification_source, f"未知({modification_source})")
        }
        parsed_records.append(record_dict)

    print("\n--- 解析完成 ---")
    return header_info, parsed_records

def export_to_txt_csv(header, records, output_filename):
    """将解析出的数据导出为 TXT (CSV 格式) 文件。"""
    # 使用 'utf-8-sig' 编码可以确保在 Windows Excel 中打开 CSV 文件时不会出现中文乱码
    with open(output_filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        # 写入文件头信息
        writer.writerow(["--- 文件头信息 ---"])
        for key, value in header.items():
            writer.writerow([key, value])
        writer.writerow([]) # 写入空行以作分隔
        
        # 写入日志记录
        writer.writerow(["--- 日志记录 ---"])
        writer.writerow(CSV_HEADERS)
        for record in records:
            writer.writerow(record.values())
    print(f"\n结果已成功导出到: {output_filename}")

def export_to_excel(header, records, output_filename):
    """将解析出的数据导出为格式化的 Excel (.xlsx) 文件。"""
    if not OPENPYXL_AVAILABLE:
        print("\n错误: 'openpyxl' 库未安装。无法导出为 Excel 文件。")
        print("请运行 'pip install openpyxl' 来安装。")
        return

    wb = openpyxl.Workbook()
    
    # 第一个工作表：文件头信息
    ws_header = wb.active
    ws_header.title = "文件头信息"
    ws_header.append(["项目", "值"])
    for key, value in header.items():
        ws_header.append([key, value])
    
    # 第二个工作表：日志记录
    ws_records = wb.create_sheet(title="日志记录")
    ws_records.append(CSV_HEADERS)

    # 将表头字体设置为粗体
    bold_font = Font(bold=True)
    for cell in ws_records[1]:
        cell.font = bold_font

    # 填充记录数据
    for record in records:
        ws_records.append(list(record.values()))

    # 自动调整所有工作表的列宽，以获得最佳可读性
    for ws in [ws_header, ws_records]:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # 设置一个合适的宽度（最大长度 + 2个字符的边距）
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_filename)
    print(f"\n结果已成功导出到: {output_filename}")

# ==============================================================================
# 主程序入口
# ==============================================================================

if __name__ == "__main__":
    try:
        # 1. 提示用户输入文件名
        input_filename = input("请输入要解析的日志文件名 (例如: reg_log_dump.txt): ")
        
        # 2. 调用核心函数解析文件
        header_info, records_list = parse_log_file(input_filename)

        # 3. 如果解析成功，则进入导出流程
        if header_info and records_list:
            while True:
                print("\n请选择输出格式:")
                print("  1: TXT (CSV 格式)")
                print("  2: Excel (.xlsx)")
                choice = input("请输入选项 (1 或 2): ")
                
                # 从输入文件名中提取不含扩展名的部分，用于构造输出文件名
                base_filename = os.path.splitext(input_filename)[0]

                if choice == '1':
                    output_filename = f"{base_filename}_parsed.txt"
                    export_to_txt_csv(header_info, records_list, output_filename)
                    break
                elif choice == '2':
                    output_filename = f"{base_filename}_parsed.xlsx"
                    export_to_excel(header_info, records_list, output_filename)
                    break
                else:
                    print("无效输入，请输入 1 或 2。")

    except KeyboardInterrupt:
        # 捕获 Ctrl+C，使用户可以优雅地中止程序
        print("\n\n操作已由用户取消。")
    except Exception as e:
        # 捕获所有其他未知异常，打印错误信息以供调试
        print(f"\n发生未知错误: {e}")