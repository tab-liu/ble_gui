"""
@File      : parse_energy_unit_file.py
@Version   : 1.0
@Author    : GitHub Copilot
@Date      : 2025/11/20
@Brief     : 解析 ESP32 项目能量单元文件（energy_file_header_t 和 energy_file_record_unit_t）。
@Details   :
  - 输入: 从设备导出的能量单元文件（二进制或十六进制字符串）。
  - 处理:
    1. 读取文件内容（支持二进制或十六进制字符串）。
    2. 解析文件头（energy_file_header_t，固定512字节）。
    3. 解析每个能量记录单元（energy_file_record_unit_t，固定长度）。
    4. 格式化输出，支持 TXT/CSV/Excel 导出。
"""

import struct
import datetime
import sys
import os
import binascii
import csv

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ===================== 常量定义（与 energy_process.h 保持一致） =====================
ENERGY_FILE_HEADER_LEN = 512
ENERGY_FILE_RECORD_UNIT_LEN = 50
ENERGY_ITEM_MAX_NUM = 15

# ===================== 结构体格式定义 =====================
# 文件头结构体格式（仅解析主要字段，详细解析可扩展）
HEADER_STRUCT_FORMAT = '<BBHIIIIIB' + f'{ENERGY_ITEM_MAX_NUM}s'  # 前面字段见 energy_file_header_t
# 能量统计结构体格式
KWH_ENERGY_U32_FORMAT = f'{ENERGY_ITEM_MAX_NUM}I'
KWH_ENERGY_U16_FORMAT = f'{ENERGY_ITEM_MAX_NUM}H'
BASE_CHARGE_PLAN_FORMAT = 'HHHhB11s'  # 见 base_charge_plan_t

# 能量记录单元结构体格式
RECORD_STRUCT_FORMAT = KWH_ENERGY_U16_FORMAT + BASE_CHARGE_PLAN_FORMAT

ENERGY_ITEM_NAMES = [
    "pv", "gr", "fd", "ac", "dc", "ap", "pa", "bc", "bd", "pl", "cc",
    "r12", "r13", "r14", "r15"
]

ENERGY_ITEM_NAMES_CN = [
    "总PV", "电网买电", "电网卖电", "AC负载", "DC负载", "交流PV", "总PV(重复)", "电池充电", "电池放电", "PV到AC负载", "车充",
    "保留12", "保留13", "保留14", "保留15"
]
CHARGE_PLAN_CN = ["馈电电价", "充电电价", "光伏预测", "充放电计划", "SOC", "保留"]

HEADER_CN_MAP = {
    "ver": "文件头协议版本",
    "unit_hour_period": "每小时记录次数",
    "unit_bytes": "数据单元大小（字节）",
    "first_utc_timestamp": "首次存储时间戳",
    "utc_timestamp": "最新数据时间戳",
    "unit_max_cnt": "最大节点数",
    "unit_real_cnt": "实际节点数",
    "unit_bias": "最新数据偏移",
    "item_total": "能量标签总数",
    "index": "能量标签索引"
}

def parse_energy_file(filepath):
    """
    解析能量单元文件，返回文件头和所有记录单元。
    """
    if not os.path.exists(filepath):
        print(f"\n错误: 文件 '{filepath}' 不存在。")
        return None, None

    with open(filepath, 'rb') as f:
        content = f.read()
    if all(chr(b) in '0123456789abcdefABCDEF \n\r\t' for b in content[:128]):
        try:
            hex_str = content.decode().replace('\n', '').replace('\r', '').replace(' ', '')
            content = binascii.unhexlify(hex_str)
        except Exception as e:
            print(f"十六进制字符串转换失败: {e}")
            return None, None

    file_size = len(content)
    print(f"\n--- 文件大小: {file_size} 字节 ---")

    if file_size < ENERGY_FILE_HEADER_LEN:
        print("错误: 文件不足一个完整文件头。")
        return None, None

    header_data = content[:ENERGY_FILE_HEADER_LEN]

    # 完整文件头结构体格式
    HEADER_STRUCT_FORMAT_FULL = (
        '<BBHIIIIIB' + f'{ENERGY_ITEM_MAX_NUM}s' +           # 基础头部
        f'{ENERGY_ITEM_MAX_NUM}I' +                          # delta_energy_all_100Wh
        f'{ENERGY_ITEM_MAX_NUM}I' +                          # delta_energy_year_100Wh
        f'{ENERGY_ITEM_MAX_NUM}I' +                          # delta_energy_month_100Wh
        f'{ENERGY_ITEM_MAX_NUM}I' +                          # delta_energy_day_100Wh
        f'{ENERGY_ITEM_MAX_NUM}H' +                          # delta_energy_unit_1Wh
        f'{ENERGY_ITEM_MAX_NUM}H' +                          # delta_energy_calc_1Wh
        'HHHhB11s'                                           # base_charge_plan_t
    )

    header_unpack = struct.unpack_from(HEADER_STRUCT_FORMAT_FULL, header_data)

    offset = 0
    header_info = {}
    header_info["ver"] = header_unpack[offset]; offset += 1
    header_info["unit_hour_period"] = header_unpack[offset]; offset += 1
    header_info["unit_bytes"] = header_unpack[offset]; offset += 1
    header_info["first_utc_timestamp"] = header_unpack[offset]; offset += 1
    header_info["utc_timestamp"] = header_unpack[offset]; offset += 1
    header_info["unit_max_cnt"] = header_unpack[offset]; offset += 1
    header_info["unit_real_cnt"] = header_unpack[offset]; offset += 1
    header_info["unit_bias"] = header_unpack[offset]; offset += 1
    header_info["item_total"] = header_unpack[offset]; offset += 1
    header_info["index"] = list(header_unpack[offset]); offset += 1
    header_info["delta_energy_all_100Wh"] = list(header_unpack[offset:offset+ENERGY_ITEM_MAX_NUM]); offset += ENERGY_ITEM_MAX_NUM
    header_info["delta_energy_year_100Wh"] = list(header_unpack[offset:offset+ENERGY_ITEM_MAX_NUM]); offset += ENERGY_ITEM_MAX_NUM
    header_info["delta_energy_month_100Wh"] = list(header_unpack[offset:offset+ENERGY_ITEM_MAX_NUM]); offset += ENERGY_ITEM_MAX_NUM
    header_info["delta_energy_day_100Wh"] = list(header_unpack[offset:offset+ENERGY_ITEM_MAX_NUM]); offset += ENERGY_ITEM_MAX_NUM
    header_info["delta_energy_unit_1Wh"] = list(header_unpack[offset:offset+ENERGY_ITEM_MAX_NUM]); offset += ENERGY_ITEM_MAX_NUM
    header_info["delta_energy_calc_1Wh"] = list(header_unpack[offset:offset+ENERGY_ITEM_MAX_NUM]); offset += ENERGY_ITEM_MAX_NUM
    base_charge_plan = {}
    base_charge_plan["price_Feedback"] = header_unpack[offset]; offset += 1
    base_charge_plan["price_Chgin"] = header_unpack[offset]; offset += 1
    base_charge_plan["Pv_forecast"] = header_unpack[offset]; offset += 1
    base_charge_plan["ac_charge_plan"] = header_unpack[offset]; offset += 1
    base_charge_plan["soc"] = header_unpack[offset]; offset += 1
    base_charge_plan["reserved"] = binascii.hexlify(header_unpack[offset]).decode(); offset += 1
    header_info["base_charge_plan"] = base_charge_plan

    print("\n--- 文件头信息（完整解析） ---")
    for k, v in header_info.items():
        if 'timestamp' in k:
            try:
                dt = datetime.datetime.fromtimestamp(v)
                print(f"{k:<25}: {v} ({dt.strftime('%Y-%m-%d %H:%M:%S')})")
            except:
                print(f"{k:<25}: {v} (无效时间戳)")
        elif k == "index":
            print(f"{k:<25}: {v} ({[ENERGY_ITEM_NAMES[i-1] if 1 <= i <= len(ENERGY_ITEM_NAMES) else 'unk' for i in v]})")
        elif k.startswith("delta_energy"):
            print(f"{k:<25}: {v}")
        elif k == "base_charge_plan":
            print(f"{k:<25}: {v}")
        else:
            print(f"{k:<25}: {v}")

    # 2. 解析所有记录单元
    records_data = content[ENERGY_FILE_HEADER_LEN:]
    num_records = len(records_data) // ENERGY_FILE_RECORD_UNIT_LEN
    print(f"\n--- 发现 {num_records} 条能量记录单元 ---")
    records = []

    for i in range(num_records):
        unit_chunk = records_data[i * ENERGY_FILE_RECORD_UNIT_LEN : (i + 1) * ENERGY_FILE_RECORD_UNIT_LEN]
        delta_energy = struct.unpack_from(KWH_ENERGY_U16_FORMAT, unit_chunk)
        charge_plan = struct.unpack_from(BASE_CHARGE_PLAN_FORMAT, unit_chunk[2*ENERGY_ITEM_MAX_NUM:])
        record = {
            "序号": i + 1,
            "delta_energy": dict(zip(ENERGY_ITEM_NAMES, delta_energy)),
            "charge_plan": {
                "price_Feedback": charge_plan[0],
                "price_Chgin": charge_plan[1],
                "Pv_forecast": charge_plan[2],
                "ac_charge_plan": charge_plan[3],
                "soc": charge_plan[4],
                "reserved": binascii.hexlify(charge_plan[5]).decode()
            }
        }
        records.append(record)

    print("\n--- 解析完成 ---")
    return header_info, records

def export_to_excel(header, records, output_filename):
    """导出为 Excel 文件，表头中文，能量增量部分明确单位和中文表述，充放电计划用结构体字符串展示，自动调整较大列宽"""
    if not OPENPYXL_AVAILABLE:
        print("\n错误: 'openpyxl' 库未安装。无法导出为 Excel 文件。")
        print("请运行 'pip install openpyxl' 来安装。")
        return

    wb = openpyxl.Workbook()
    ws_header = wb.active
    ws_header.title = "文件头信息"
    ws_header.append(["项目", "值", "说明"])

    # 基础字段
    for key in [
        "ver", "unit_hour_period", "unit_bytes", "first_utc_timestamp", "utc_timestamp",
        "unit_max_cnt", "unit_real_cnt", "unit_bias", "item_total", "index"
    ]:
        value = header.get(key)
        cn = HEADER_CN_MAP.get(key, key)
        if key in ("first_utc_timestamp", "utc_timestamp"):
            try:
                dt = datetime.datetime.fromtimestamp(value)
                ws_header.append([cn, f"{value}", dt.strftime("%Y-%m-%d %H:%M:%S")])
            except:
                ws_header.append([cn, f"{value}", "无效时间"])
        elif key == "index":
            idx_str = ",".join(str(v) for v in value)
            name_str = ",".join([ENERGY_ITEM_NAMES_CN[i-1] if 1 <= i <= len(ENERGY_ITEM_NAMES_CN) else "未知" for i in value])
            ws_header.append([cn, idx_str, name_str])
        else:
            ws_header.append([cn, value, ""])

    # 能量增量部分，明确单位和中文表述
    energy_sections = [
        ("delta_energy_all_100Wh", "总能量增量", "各路累计总能量，单位：0.1kWh"),
        ("delta_energy_year_100Wh", "年能量增量", "各路本年累计能量，单位：0.1kWh"),
        ("delta_energy_month_100Wh", "月能量增量", "各路本月累计能量，单位：0.1kWh"),
        ("delta_energy_day_100Wh", "日能量增量", "各路本日累计能量，单位：0.1kWh"),
        ("delta_energy_unit_1Wh", "节点能量增量", "各路节点能量，单位：1Wh"),
        ("delta_energy_calc_1Wh", "计算能量增量", "各路计算能量，单位：1Wh"),
    ]
    for sec_key, sec_cn, sec_desc in energy_sections:
        arr = header.get(sec_key, [])
        arr_str = ", ".join(f"{ENERGY_ITEM_NAMES_CN[i]}={arr[i]}" for i in range(len(arr)))
        ws_header.append([sec_cn, arr_str, sec_desc])

    # 充放电计划结构体
    cp = header.get("base_charge_plan", {})
    cp_str = (
        f'base_charge_plan_t('
        f'馈电电价={cp.get("price_Feedback", "")}, '
        f'充电电价={cp.get("price_Chgin", "")}, '
        f'光伏预测={cp.get("Pv_forecast", "")}, '
        f'充放电计划={cp.get("ac_charge_plan", "")}, '
        f'SOC={cp.get("soc", "")}, '
        f'保留={cp.get("reserved", "")})'
    )
    ws_header.append(["充放电计划结构体", cp_str, "当前充放电计划及相关信息"])

    ws_records = wb.create_sheet(title="能量记录单元")
    ws_records.append(["序号", "时间"] + ENERGY_ITEM_NAMES_CN + ["充放电计划结构体"])
    bold_font = Font(bold=True)
    for cell in ws_records[1]:
        cell.font = bold_font

    # 时间推算
    try:
        t_first = header["first_utc_timestamp"]
        t_last = header["utc_timestamp"]
        count = header["unit_real_cnt"]
        if count > 1:
            interval = (t_last - t_first) // (count - 1)
        else:
            interval = 0
    except Exception:
        t_first = t_last = interval = 0

    for i, record in enumerate(records):
        node_time = t_first + i * interval if interval else t_last
        try:
            node_time_str = datetime.datetime.fromtimestamp(node_time).strftime("%Y-%m-%d %H:%M:%S")
        except:
            node_time_str = "无效时间"
        row = [record["序号"], node_time_str]
        row += [record["delta_energy"].get(name, "") for name in ENERGY_ITEM_NAMES]
        cp = record["charge_plan"]
        cp_str = (
            f'base_charge_plan_t('
            f'馈电电价={cp["price_Feedback"]}, '
            f'充电电价={cp["price_Chgin"]}, '
            f'光伏预测={cp["Pv_forecast"]}, '
            f'充放电计划={cp["ac_charge_plan"]}, '
            f'SOC={cp["soc"]}, '
            f'保留={cp["reserved"]})'
        )
        row.append(cp_str)
        ws_records.append(row)

    # 自动调整较大列宽，满足中文描述
    for ws in [ws_header, ws_records]:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 8  # 增加间距

    wb.save(output_filename)
    print(f"\n结果已成功导出到: {output_filename}")

if __name__ == "__main__":
    try:
        input_filename = input("请输入能量单元文件名 (如: energy_unit_dump.bin): ")
        header_info, records_list = parse_energy_file(input_filename)
        if header_info and records_list:
            while True:
                print("\n请选择输出格式:")
                print("  1: TXT (CSV 格式)")
                print("  2: Excel (.xlsx)")
                choice = input("请输入选项 (1 或 2): ")
                base_filename = os.path.splitext(input_filename)[0]
                if choice == '1':
                    output_filename = f"{base_filename}_parsed.txt"
                    export_to_txt_csv(header_info, records_list, output_filename)
                    break
                elif choice == '2':
                    output_filename = f"{base_filename}_parsed.xlsx"
                    export_to_excel(header_info, records_list, output_filename)
                    break
                else:
                    print("无效输入，请输入 1 或 2。")
    except KeyboardInterrupt:
        print("\n\n操作已由用户取消。")
    except Exception as e:
        print(f"\n发生未知错误: {e}")