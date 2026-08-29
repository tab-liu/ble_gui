"""
@File      : parse_energy_year_record.py
@Version   : 1.0
@Author    : GitHub Copilot
@Date      : 2025/11/20
@Brief     : 解析 ESP32 项目年度能量统计文件（energy_file_year_record_t）。
@Details   :
  - 输入: 由 energy_file_year_record_t 结构体序列化生成的二进制文件（每条128字节）。
  - 处理:
    1. 读取文件内容（支持二进制或十六进制字符串）。
    2. 逐条解析年度能量统计结构体。
    3. 格式化输出，支持 TXT/CSV/Excel 导出。
"""

import struct
import datetime
import sys
import os
import binascii
import csv

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

ENERGY_LONG_TERM_DATA_MAX_LEN = 128

# 结构体格式定义（与 energy_file_year_record_t 完全一致）
YEAR_RECORD_FORMAT = '<HHI12I31H H'  # type, year, total_energy, mouth_energy[12], day_enengy[31], crc16

ENERGY_TYPE_MAP = {
    1: "PV", 2: "GRID_IN", 3: "GRID_OUT", 4: "AC_LOAD", 5: "DC_LOAD",
    6: "AC_PV", 7: "PV_AGAIN", 8: "BAT_CHARGE", 9: "BAT_DISCHARGE",
    10: "PV_TO_AC", 11: "CAR_CHARGE"
}

def parse_year_record_file(filepath):
    """
    解析年度能量统计文件，返回所有记录列表。
    """
    if not os.path.exists(filepath):
        print(f"\n错误: 文件 '{filepath}' 不存在。")
        return None

    with open(filepath, 'rb') as f:
        content = f.read()
    # 支持十六进制字符串文件自动转换
    if all(chr(b) in '0123456789abcdefABCDEF \n\r\t' for b in content[:128]):
        try:
            hex_str = content.decode().replace('\n', '').replace('\r', '').replace(' ', '')
            content = binascii.unhexlify(hex_str)
        except Exception as e:
            print(f"十六进制字符串转换失败: {e}")
            return None

    file_size = len(content)
    num_records = file_size // ENERGY_LONG_TERM_DATA_MAX_LEN
    print(f"\n--- 文件大小: {file_size} 字节，共 {num_records} 条年度能量记录 ---")
    records = []

    for i in range(num_records):
        chunk = content[i * ENERGY_LONG_TERM_DATA_MAX_LEN : (i + 1) * ENERGY_LONG_TERM_DATA_MAX_LEN]
        unpacked = struct.unpack_from(YEAR_RECORD_FORMAT, chunk)
        record = {
            "index": i + 1,
            "type": unpacked[0],
            "type_name": ENERGY_TYPE_MAP.get(unpacked[0], f"未知({unpacked[0]})"),
            "year": unpacked[1],
            "total_energy": unpacked[2],
            "mouth_energy": list(unpacked[3:15]),  # 12个月
            "day_enengy": list(unpacked[15:46]),   # 31天
            "crc16": unpacked[46]
        }
        records.append(record)
    print("\n--- 解析完成 ---")
    return records

def export_to_excel(records, output_filename):
    """导出为 Excel 文件，表头注明单位为0.1kWh，自动调整较大列宽"""
    if not OPENPYXL_AVAILABLE:
        print("\n错误: 'openpyxl' 库未安装。无法导出为 Excel 文件。")
        print("请运行 'pip install openpyxl' 来安装。")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "年度能量统计"
    # 中文表头，单位注明为0.1kWh
    ws.append([
        "序号", "能量类型编号", "能量类型名称", "统计年份", "年度总能量（0.1kWh）",
        "1月能量（0.1kWh）", "2月能量（0.1kWh）", "3月能量（0.1kWh）", "4月能量（0.1kWh）", "5月能量（0.1kWh）", "6月能量（0.1kWh）",
        "7月能量（0.1kWh）", "8月能量（0.1kWh）", "9月能量（0.1kWh）", "10月能量（0.1kWh）", "11月能量（0.1kWh）", "12月能量（0.1kWh）",
        "1日能量（0.1kWh）", "2日能量（0.1kWh）", "3日能量（0.1kWh）", "4日能量（0.1kWh）", "5日能量（0.1kWh）", "6日能量（0.1kWh）", "7日能量（0.1kWh）",
        "8日能量（0.1kWh）", "9日能量（0.1kWh）", "10日能量（0.1kWh）", "11日能量（0.1kWh）", "12日能量（0.1kWh）", "13日能量（0.1kWh）", "14日能量（0.1kWh）",
        "15日能量（0.1kWh）", "16日能量（0.1kWh）", "17日能量（0.1kWh）", "18日能量（0.1kWh）", "19日能量（0.1kWh）", "20日能量（0.1kWh）", "21日能量（0.1kWh）",
        "22日能量（0.1kWh）", "23日能量（0.1kWh）", "24日能量（0.1kWh）", "25日能量（0.1kWh）", "26日能量（0.1kWh）", "27日能量（0.1kWh）", "28日能量（0.1kWh）",
        "29日能量（0.1kWh）", "30日能量（0.1kWh）", "31日能量（0.1kWh）", "CRC16校验值"
    ])
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for record in records:
        row = [
            record["index"], record["type"], record["type_name"], record["year"], record["total_energy"]
        ] + record["mouth_energy"] + record["day_enengy"] + [record["crc16"]]
        ws.append(row)

    # 自动调整较大列宽
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 8  # 增加间距

    wb.save(output_filename)
    print(f"\n结果已成功导出到: {output_filename}")

def export_to_txt_csv(records, output_filename):
    """导出为 TXT (CSV 格式) 文件，表头注明单位为0.1kWh"""
    with open(output_filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            "序号", "type", "type_name", "year", "total_energy(0.1kWh)",
            *(f"mouth_energy[{i+1}](0.1kWh)" for i in range(12)),
            *(f"day_enengy[{i+1}](0.1kWh)" for i in range(31)),
            "crc16"
        ])
        for record in records:
            row = [
                record["index"], record["type"], record["type_name"], record["year"], record["total_energy"]
            ] + record["mouth_energy"] + record["day_enengy"] + [record["crc16"]]
            writer.writerow(row)
    print(f"\n结果已成功导出到: {output_filename}")

if __name__ == "__main__":
    try:
        input_filename = input("请输入年度能量统计文件名 (如: energy_year_record.bin): ")
        records = parse_year_record_file(input_filename)
        if records:
            while True:
                print("\n请选择输出格式:")
                print("  1: TXT (CSV 格式)")
                print("  2: Excel (.xlsx)")
                choice = input("请输入选项 (1 或 2): ")
                base_filename = os.path.splitext(input_filename)[0]
                if choice == '1':
                    output_filename = f"{base_filename}_parsed.txt"
                    export_to_txt_csv(records, output_filename)
                    break
                elif choice == '2':
                    output_filename = f"{base_filename}_parsed.xlsx"
                    export_to_excel(records, output_filename)
                    break
                else:
                    print("无效输入，请输入 1 或 2。")
    except KeyboardInterrupt:
        print("\n\n操作已由用户取消。")
    except Exception as e:
        print(f"\n发生未知错误: {e}")