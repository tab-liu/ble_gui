"""
@File      : parse_dev_access_log.py
@Version   : 1.2
@Author    : GitHub Copilot
@Date      : 2025/10/13
@Brief     : 用于解析 ESP32 项目中 dev_access_log 模块生成的日志文件。
@Details   :
  - 输入: 一个 .txt 文件，其内容为从设备读取的二进制日志数据的十六进制字符串表示。
  - 处理:
    1. 读取文件中的十六进制字符串。
    2. 将其转换回原始的二进制字节流。
    3. 根据 C 头文件 (dev_access_log.h) 中定义的结构体，解析文件头和每一条日志记录。
    4. 将解析结果进行格式化，使其易于人类阅读。
  - 输出:
    - 在控制台打印解析结果。
    - 可选择将结果导出为结构化的 TXT (CSV格式) 或 Excel (.xlsx) 文件。
"""

# ==============================================================================
# 依赖库导入
# ==============================================================================
import struct         # 用于处理二进制数据，执行打包和解包操作
import datetime       # 用于将 UNIX 时间戳转换为可读的日期时间格式
import sys            # 用于处理程序退出
import os             # 用于处理文件路径和文件名
import csv            # 用于生成 CSV 格式的 TXT 文件
import binascii       # 用于在二进制和 ASCII 编码的十六进制字符串之间进行转换

# 尝试导入 openpyxl，如果失败则标记为不可用，以便在后续流程中优雅地处理
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==============================================================================
# 全局常量定义 (与 C 头文件 dev_access_log.h 严格对应)
# ==============================================================================

# 定义文件头和每条记录的固定长度（字节）
DEV_ACCESS_FILE_HEADER_LEN = 20
DEV_ACCESS_LOG_LEN = 40

# 定义二进制解包格式字符串
# '<' 表示小端字节序 (Little-endian)
HEADER_STRUCT_FORMAT = '<HHH'          # 对应 device_event_file_header_t.s (uint16_t * 3 = 6字节)
RECORD_BASE_STRUCT_FORMAT = '<IBBBBB'  # 对应记录结构体的前9个字节
SN_INFO_FORMAT = '<12sQ'               # 对应联合体中的 dev_type (char[12]) 和 dev_sn (uint64_t)

# --- 枚举值到字符串的映射 ---
RECORD_TYPE_MAP = {
    0: "通用整机设备（未知来源）",
    1: "整机设备（串口总线）",
    2: "整机设备（CAN总线）",
    3: "整机设备（蓝牙数据广播）",
    4: "整机设备（蓝牙Mesh）",
    5: "整机设备（WIFI Mesh）",
    6: "整机设备（WIFI UDP）",
    7: "整机设备（Sub1G）",
    8: "整机设备（CAN总线-2）",
    9: "整机设备（RS485总线）",
    0xFC: "MCU模块（CAN总线架构系统）",
    0xFD: "MCU模块（串口总线架构系统）",
    0xFE: "MCU模块（RS485总线架构系统）",
    0xFF: "通用MCU模块（未知来源）",
}

OPERATION_ATTRIBUTE_MAP = {
    1: "自动识别增加",
    2: "设备超时离线",
    3: "外部强制绑定",
    4: "外部强制解绑",
}

INFO_TYPE_MAP = {
    0: "标准SN",
    1: "蓝牙广播名",
}

# 定义导出文件的列标题
CSV_HEADERS = [
    "记录序号", "时间戳", "可读时间", "记录类型", "上级地址", "本机地址",
    "操作属性", "信息类型", "设备标识符"
]

# ==============================================================================
# 核心功能函数
# ==============================================================================

def parse_log_file(filepath):
    """
    解析包含十六进制字符串的日志文件，并返回文件头信息和记录列表。

    @param  filepath: 日志文件的路径。
    @return: 一个元组 (header_info, parsed_records)，如果失败则返回 (None, None)。
    """
    if not os.path.exists(filepath):
        print(f"\n错误: 文件 '{filepath}' 不存在或路径不正确。")
        return None, None

    # 1. 读取文件内容 (以文本模式)
    with open(filepath, 'r') as f:
        hex_string = f.read().strip()

    # 2. 【关键步骤】将十六进制字符串转换为二进制数据
    try:
        hex_string = "".join(hex_string.split())
        binary_content = binascii.unhexlify(hex_string)
    except (binascii.Error, TypeError) as e:
        print(f"\n错误: 文件内容不是有效的十六进制字符串。 {e}")
        return None, None

    file_size = len(binary_content)
    print(f"\n--- 开始解析文件: '{filepath}' (有效数据大小: {file_size} 字节) ---")

    # 3. 验证并解析文件头
    if file_size < DEV_ACCESS_FILE_HEADER_LEN:
        print(f"错误: 数据大小 ({file_size}字节) 不足以包含一个完整的文件头 ({DEV_ACCESS_FILE_HEADER_LEN}字节)。")
        return None, None

    header_data = binary_content[:DEV_ACCESS_FILE_HEADER_LEN]
    max_records, current_records, write_index = struct.unpack_from(HEADER_STRUCT_FORMAT, header_data)
    header_info = {
        "max_records": max_records,
        "current_records": current_records,
        "write_index": write_index
    }
    print("\n--- 文件头信息 ---")
    for key, value in header_info.items():
        print(f"{key:<20}: {value}")
    print("-" * 20)

    # 4. 循环解析所有日志记录
    records_data = binary_content[DEV_ACCESS_FILE_HEADER_LEN:]
    num_records_in_file = len(records_data) // DEV_ACCESS_LOG_LEN
    parsed_records = []

    print(f"\n--- 发现 {num_records_in_file} 条日志记录 ---")

    for i in range(num_records_in_file):
        record_chunk = records_data[i * DEV_ACCESS_LOG_LEN : (i + 1) * DEV_ACCESS_LOG_LEN]
        
        base_part_size = struct.calcsize(RECORD_BASE_STRUCT_FORMAT)
        (timestamp, record_type, parent_address, local_address, 
         operation_attribute, info_type) = struct.unpack_from(RECORD_BASE_STRUCT_FORMAT, record_chunk)

        identifier_chunk = record_chunk[base_part_size : base_part_size + 25]
        
        identifier_str = ""
        if info_type == 0: # 标准SN
            try:
                dev_type_bytes, dev_sn = struct.unpack_from(SN_INFO_FORMAT, identifier_chunk)
                dev_type_str = dev_type_bytes.decode('utf-8', errors='ignore').split('\x00', 1)[0]
                identifier_str = f"类型: {dev_type_str}, SN: {dev_sn}"
            except struct.error:
                identifier_str = "SN信息解析失败"
        elif info_type == 1: # 蓝牙广播名
            ble_name_hex = identifier_chunk.hex()
            identifier_str = f"原始Hex: {ble_name_hex}"
        else:
            identifier_str = f"未知信息类型({info_type})"

        try:
            dt_object = datetime.datetime.fromtimestamp(timestamp)
            time_str = dt_object.strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, OSError):
            time_str = "无效时间戳"

        record_dict = {
            "记录序号": i + 1,
            "时间戳": timestamp,
            "可读时间": time_str,
            "记录类型": RECORD_TYPE_MAP.get(record_type, f"未知({record_type})"),
            "上级地址": hex(parent_address),
            "本机地址": hex(local_address),
            "操作属性": OPERATION_ATTRIBUTE_MAP.get(operation_attribute, f"未知({operation_attribute})"),
            "信息类型": INFO_TYPE_MAP.get(info_type, f"未知({info_type})"),
            "设备标识符": identifier_str
        }
        parsed_records.append(record_dict)

    print("\n--- 解析完成 ---")
    return header_info, parsed_records

def export_to_txt_csv(header, records, output_filename):
    """将解析出的数据导出为 TXT (CSV 格式) 文件。"""
    with open(output_filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        writer.writerow(["--- 文件头信息 ---"])
        for key, value in header.items():
            writer.writerow([key, value])
        writer.writerow([])
        
        writer.writerow(["--- 日志记录 ---"])
        writer.writerow(CSV_HEADERS)
        for record in records:
            writer.writerow(record.values())
    print(f"\n结果已成功导出到: {output_filename}")

def export_to_excel(header, records, output_filename):
    """将解析出的数据导出为格式化的 Excel (.xlsx) 文件。"""
    if not OPENPYXL_AVAILABLE:
        print("\n错误: 'openpyxl' 库未安装。无法导出为 Excel 文件。")
        print("请运行 'pip install openpyxl' 来安装。")
        return

    wb = openpyxl.Workbook()
    
    ws_header = wb.active
    ws_header.title = "文件头信息"
    ws_header.append(["项目", "值"])
    for key, value in header.items():
        ws_header.append([key, value])
    
    ws_records = wb.create_sheet(title="日志记录")
    ws_records.append(CSV_HEADERS)

    bold_font = Font(bold=True)
    for cell in ws_records[1]:
        cell.font = bold_font

    for record in records:
        ws_records.append(list(record.values()))

    for ws in [ws_header, ws_records]:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_filename)
    print(f"\n结果已成功导出到: {output_filename}")

# ==============================================================================
# 主程序入口
# ==============================================================================

if __name__ == "__main__":
    try:
        input_filename = input("请输入要解析的日志文件名 (例如: dev_access_log_dump.txt): ")
        
        header_info, records_list = parse_log_file(input_filename)

        if header_info and records_list:
            while True:
                print("\n请选择输出格式:")
                print("  1: TXT (CSV 格式)")
                print("  2: Excel (.xlsx)")
                choice = input("请输入选项 (1 或 2): ")
                
                base_filename = os.path.splitext(input_filename)[0]

                if choice == '1':
                    output_filename = f"{base_filename}_parsed.txt"
                    export_to_txt_csv(header_info, records_list, output_filename)
                    break
                elif choice == '2':
                    output_filename = f"{base_filename}_parsed.xlsx"
                    export_to_excel(header_info, records_list, output_filename)
                    break
                else:
                    print("无效输入，请输入 1 或 2。")

    except KeyboardInterrupt:
        print("\n\n操作已由用户取消。")
    except Exception as e:
        print(f"\n发生未知错误: {e}")